"""
Simplified Patient Medication Registry Builder
===============================================

Focused extraction pipeline using Med7 and medspaCy for seizure-related clinical notes.

Process:
1. Load seizure_context.txt files for each patient
2. Use Med7 to extract: drugs, dosages, frequencies, routes
3. Use medspaCy ConText to extract: seizure status and symptoms
4. Build structured registry with key medication and seizure information
5. Save one parquet file per patient

Output Columns:
- patient_id
- note_date
- medication
- medication_dosage
- intake_times_per_day
- medication_effectiveness
- seizure_status (positive/negative/unknown)
- seizure_symptoms
- note_text
"""

import os
import re
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List, Tuple
from collections import defaultdict

import pandas as pd
import numpy as np
import warnings

warnings.filterwarnings('ignore')

# Import preprocessing
from preprocess import preprocess_clinical_text

# Import medication/seizure extraction helpers (moved out for readability)
from medi_info_extract import (
    set_nlp_models,
    extract_med7_entities,
    link_drug_to_attributes,
    detect_seizure_info_with_context,
    infer_effectiveness,
    extract_patient_id_from_folder,
    parse_frequency_to_per_day,
    parse_dosage_text,
    count_medication_mentions,
)

# -------------------------------
# CONFIG
# -------------------------------
def _get_project_root() -> Path:
    """
    Resolve project root without hard-coding a machine-specific path.

    Priority:
    1) Environment variable `LLM_RAG_ROOT` (lets you point to data elsewhere)
    2) Directory containing this script (portable, safe to commit)
    """
    env_root = os.environ.get("LLM_RAG_ROOT")
    if env_root:
        return Path(env_root).expanduser().resolve()
    return Path(__file__).resolve().parent


ROOT = _get_project_root()
PATIENT_FILES_DIR = ROOT / "patient_files"
OUTPUT_DIR = ROOT / "patient_registries"  # Output parquet files per patient
NOTES_DIR = ROOT / "patient_notes"  # Saved notes per patient

# Preprocessing
USE_ENHANCED_PREPROCESSING = True  # Apply comprehensive text cleaning

# Excel Export
EXPORT_TO_EXCEL = True  # Set to True to export registry to Excel files for visualization

# Check if openpyxl is available for Excel export
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    get_column_letter = None
    if EXPORT_TO_EXCEL:
        print("[WARN] openpyxl not installed. Excel export disabled.")
        print("       Install with: pip install openpyxl")
        EXPORT_TO_EXCEL = False

# -------------------------------
# LOAD MED7 AND MEDSPACY
# -------------------------------
print("="*80)
print("PATIENT MEDICATION REGISTRY BUILDER")
print("="*80)

# Load + configure NLP models (Med7 + medspaCy) via helper module
set_nlp_models()
print("\n[OK] All NLP models loaded successfully!")


# -------------------------------
# PIPELINE FUNCTIONS (READ FIRST)
# -------------------------------

def process_all_patients() -> None:
    """Main processing function."""
    
    # Create output directories
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    NOTES_DIR.mkdir(parents=True, exist_ok=True)
    
    # Find all patient folders
    patient_folders = [
        f for f in PATIENT_FILES_DIR.iterdir()
        if f.is_dir() and f.name.startswith("Patient_")
    ]
    
    if not patient_folders:
        print(f"[ERROR] No patient folders found in {PATIENT_FILES_DIR}")
        return
    
    print(f"\n[INFO] Found {len(patient_folders)} patient folders")
    print(f"[INFO] Output directory: {OUTPUT_DIR}")
    print(f"[INFO] Notes directory: {NOTES_DIR}")
    
    # Store all patient dataframes for combined file
    all_patient_dfs = []
    
    # Process each patient
    for patient_folder in patient_folders:
        df_patient = process_patient_file(patient_folder)
        
        if df_patient.empty:
            print(f"   [WARN] No data extracted for this patient")
            continue
        
        # Add to combined list
        all_patient_dfs.append(df_patient)
        
        patient_id = df_patient['patient_id'].iloc[0]
        
        print(f"\n   [OK] Processed {len(df_patient)} medication records for Patient {patient_id}")
        print(f"        Notes saved to: {NOTES_DIR / patient_id}/")
        
        # # Save patient-specific parquet file (COMMENTED OUT - only saving combined file)
        # output_file = OUTPUT_DIR / f"{patient_id}_medication_registry.parquet"
        # df_patient.to_parquet(output_file, index=False)
        # 
        # # Also save CSV for easy viewing
        # output_csv = OUTPUT_DIR / f"{patient_id}_medication_registry.csv"
        # df_patient.to_csv(output_csv, index=False, encoding='utf-8')
        # 
        # # Export to Excel if enabled
        # excel_exported = False
        # if EXPORT_TO_EXCEL:
        #     excel_exported = export_to_excel(df_patient, output_file, sheet_name="Medication Registry")
        # 
        # print(f"\n   [OK] Saved {len(df_patient)} medication records")
        # print(f"        Parquet: {output_file}")
        # print(f"        CSV: {output_csv}")
        # if excel_exported:
        #     excel_path = output_file.with_suffix('.xlsx')
        #     print(f"        Excel: {excel_path}")
        # print(f"        Notes: {NOTES_DIR / patient_id}/")
        
        # Show summary stats
        print(f"\n   Summary for Patient {patient_id}:")
        print(f"     Total records: {len(df_patient)}")
        print(f"     Unique medications: {df_patient['medication'].nunique()}")
        print(f"     Date range: {df_patient['note_date'].min()} to {df_patient['note_date'].max()}")
        print(f"     Seizure positive notes: {(df_patient['seizure_status'] == 'positive').sum()}")
        print(f"     Seizure negative notes: {(df_patient['seizure_status'] == 'negative').sum()}")
        
        # Top medications
        if len(df_patient) > 0:
            top_meds = df_patient['medication'].value_counts().head(3)
            print(f"     Top medications:")
            for med, count in top_meds.items():
                print(f"       - {med}: {count} records")
    
    # Create combined summary file
    print(f"\n{'='*80}")
    print("Creating Combined Summary")
    print(f"{'='*80}")
    
    if all_patient_dfs:
        df_combined = pd.concat(all_patient_dfs, ignore_index=True)
        
        combined_file = OUTPUT_DIR / "all_patients_combined.parquet"
        df_combined.to_parquet(combined_file, index=False)
        
        combined_csv = OUTPUT_DIR / "all_patients_combined.csv"
        df_combined.to_csv(combined_csv, index=False, encoding='utf-8')
        
        # Export combined file to Excel if enabled
        excel_exported = False
        if EXPORT_TO_EXCEL:
            excel_exported = export_to_excel(df_combined, combined_file, sheet_name="All Patients Registry")
        
        print(f"\n[OK] Combined registry saved:")
        print(f"     {combined_file}")
        print(f"     {combined_csv}")
        if excel_exported:
            combined_excel = combined_file.with_suffix('.xlsx')
            print(f"     {combined_excel}")
        print(f"     Total records: {len(df_combined)}")
        print(f"     Total patients: {df_combined['patient_id'].nunique()}")
        print(f"     Unique medications: {df_combined['medication'].nunique()}")
        
        # Overall statistics
        print(f"\n[INFO] Overall Statistics:")
        print(f"     Total medication records: {len(df_combined)}")
        print(f"     Records with dosage: {df_combined['medication_dosage'].notna().sum()}")
        print(f"     Records with frequency: {df_combined['intake_times_per_day'].notna().sum()}")
        print(f"     Effective medications: {(df_combined['medication_effectiveness'] == 'successful').sum()}")
        print(f"     Refractory medications: {(df_combined['medication_effectiveness'] == 'refractory').sum()}")
        
        print(f"\n     Top 10 medications across all patients:")
        top_meds_all = df_combined['medication'].value_counts().head(10)
        for med, count in top_meds_all.items():
            print(f"       {count:3d} records - {med}")


def process_patient_file(patient_folder: Path) -> pd.DataFrame:
    """
    Process seizure_context.txt file for a single patient.
    Returns DataFrame with medication registry.
    """
    patient_id = extract_patient_id_from_folder(patient_folder)
    if not patient_id:
        print(f"[WARN] Could not extract patient ID from: {patient_folder}")
        return pd.DataFrame()
    
    print(f"\n{'='*80}")
    print(f"Processing Patient {patient_id}")
    print(f"{'='*80}")
    
    # Find seizure context file
    seizure_file = patient_folder / f"{patient_id}_seizure_context.txt"
    
    if not seizure_file.exists():
        print(f"[WARN] Seizure context file not found: {seizure_file}")
        return pd.DataFrame()
    
    print(f"[INFO] Reading: {seizure_file.name}")
    
    # Split file into individual notes
    notes = split_seizure_file_into_notes(seizure_file)
    print(f"[INFO] Found {len(notes)} dated notes")
    
    # Save notes to patient notes directory
    patient_notes_dir = NOTES_DIR / patient_id
    patient_notes_dir.mkdir(parents=True, exist_ok=True)
    
    # Process each note
    all_records = []
    
    for note_idx, note in enumerate(notes, 1):
        note_date = note['date']
        note_text = note['text']
        
        if not note_date:
            print(f"   [SKIP] Note {note_idx} has no date")
            continue
        
        print(f"   [{note_idx}/{len(notes)}] Processing note from {note_date}...")
        
        # Save note text
        note_filename = f"{note_date}_{patient_id}_seizure_note.txt"
        note_path = patient_notes_dir / note_filename
        try:
            note_path.write_text(note_text, encoding='utf-8')
        except Exception as e:
            print(f"        [WARN] Could not save note: {e}")
        
        # Extract entities with Med7
        entities = extract_med7_entities(note_text)
        
        # Link drugs to their attributes
        medications = link_drug_to_attributes(entities, note_text)
        
        # Detect seizure status and symptoms for this note
        seizure_status, seizure_symptoms = detect_seizure_info_with_context(note_text)
        
        print(f"        Found {len(medications)} medications, seizure status: {seizure_status}")
        
        # Build record for each medication
        note_id = f"{patient_id}_{note_date}_{note_idx:03d}"
        
        # Generate ONE event per patient per note (not per medication)
        event_id = f"{patient_id}_{note_date}_{note_idx:03d}"
        
        for med_idx, med in enumerate(medications, 1):
            drug_name = med['drug'].lower().strip()
            
            # Parse dosage
            dosage = parse_dosage_text(med['strength'])
            
            # Parse frequency to times per day
            intake_per_day = parse_frequency_to_per_day(med['frequency'])
            
            # Count mentions
            mention_count = count_medication_mentions(note_text, drug_name)
            
            # Infer effectiveness from context
            effectiveness = infer_effectiveness(med['context'])
            
            # Generate unique row_id for each row (first column)
            # Format: patient_id-row_number
            row_num = len(all_records) + 1
            row_id = f"{patient_id}-{row_num}"
            
            record = {
                'row_id': row_id,  # Unique ID for each row
                # 'event_id': event_id,  # Same event_id for all meds in this note
                # 'event_type': 'clinical_note',  # Represents the entire clinical note
                'patient_id': patient_id,
                'note_date': note_date,
                'note_id': note_id,
                'source_text_file': str(note_path),
                'source_date_key': note_date,
                
                'medication': drug_name,
                'medication_dosage': dosage,
                'intake_times_per_day': intake_per_day,   # derived from frequency
                'drug_mention_count': mention_count,       # distinct from intake_times_per_day
                'medication_effectiveness': effectiveness,
                
                'seizure_status': seizure_status,
                'seizure_symptoms': seizure_symptoms,
                
                'route': med['route'],
                'form': med['form'],
                'note_text': note_text,
                'extraction_context': med['context'],
            }
            
            all_records.append(record)
    
    # Create DataFrame
    df = pd.DataFrame(all_records)
    
    # Dtype hygiene and cleanup
    if not df.empty:
        # Drop rows with missing medication
        df = df.dropna(subset=['medication'])
        
        # Ensure intake_times_per_day is numeric
        if 'intake_times_per_day' in df:
            df['intake_times_per_day'] = pd.to_numeric(df['intake_times_per_day'], errors='coerce')
    
    return df


# -------------------------------
# NOTE + EXPORT HELPERS
# -------------------------------

def split_seizure_file_into_notes(file_path: Path) -> List[Dict]:
    """
    Split seizure_context.txt file into individual dated notes.
    Returns list of {'date': ..., 'text': ...}
    """
    try:
        # Read file
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except UnicodeDecodeError:
        with open(file_path, 'r', encoding='cp1252') as f:
            content = f.read()
    
    # Apply preprocessing if enabled
    if USE_ENHANCED_PREPROCESSING:
        content = preprocess_clinical_text(content)
    
    # Look for date headers: "FILE: ... | DATE: YYYYMMDD"
    date_pattern = re.compile(
        r"^FILE:\s*(?P<fname>.+?)\s*\|\s*DATE:\s*(?P<date>\d{8})\s*$",
        flags=re.IGNORECASE | re.MULTILINE
    )
    
    matches = list(date_pattern.finditer(content))
    
    if not matches:
        # No date headers found - treat as single note
        return [{'date': None, 'text': content}]
    
    notes = []
    for i, match in enumerate(matches):
        start_pos = match.start()
        end_pos = matches[i+1].start() if i+1 < len(matches) else len(content)
        
        date_str = match.group('date')
        # Convert YYYYMMDD to YYYY-MM-DD
        try:
            date_obj = datetime.strptime(date_str, '%Y%m%d')
            note_date = date_obj.strftime('%Y-%m-%d')
        except:
            note_date = None
        
        # Extract block and exclude header lines
        block = content[start_pos:end_pos].strip()
        block_lines = block.splitlines()
        
        # Skip the header line "FILE: ... | DATE: ..." and following bar lines
        start_line = 1
        while start_line < len(block_lines) and re.match(r'^\s*=+\s*$', block_lines[start_line]):
            start_line += 1
        
        note_text = "\n".join(block_lines[start_line:]).strip()
        
        if note_text and note_date:
            notes.append({
                'date': note_date,
                'text': note_text
            })
    
    return notes


def export_to_excel(df: pd.DataFrame, output_path: Path, sheet_name: str = "Registry") -> bool:
    """
    Export DataFrame to Excel with auto-formatted columns.
    
    Args:
        df: DataFrame to export
        output_path: Path to save Excel file (e.g., "patient_registry.xlsx")
        sheet_name: Name of the Excel sheet
    
    Returns:
        True if successful, False otherwise
    """
    if not EXPORT_TO_EXCEL or not HAS_OPENPYXL:
        return False
    
    if df.empty:
        print(f"        [SKIP] No data to export to Excel")
        return False
    
    try:
        # Create Excel file
        excel_path = output_path.with_suffix('.xlsx')
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Auto-adjust column widths
            for col_idx, column in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in column:
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                
                # Set width (cap at 50 for readability)
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return True
    except Exception as e:
        print(f"        [WARN] Excel export failed: {e}")
        return False


# -------------------------------
# EXTRACTION HELPERS
# -------------------------------

# Helper functions moved to `medi_info_extract.py`


# -------------------------------
# MAIN
# -------------------------------

def main():
    """Main entry point."""
    # Process overview:
    # - Confirm run configuration (preprocessing, excel export)
    # - Process all patients under `PATIENT_FILES_DIR`
    # - For each patient: split into dated notes -> extract meds/seizure status -> build registry rows
    # - Save combined outputs under `OUTPUT_DIR` and note text files under `NOTES_DIR`
    start_time = time.time()
    
    print(f"\n[INFO] Enhanced preprocessing: {'ENABLED' if USE_ENHANCED_PREPROCESSING else 'DISABLED'}")
    print(f"[INFO] Excel export: {'ENABLED' if EXPORT_TO_EXCEL else 'DISABLED'}")
    print(f"[INFO] Starting processing at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    try:
        process_all_patients()
    except Exception as e:
        print(f"\n[ERROR] Processing failed: {e}")
        traceback.print_exc()
        return
    
    elapsed = time.time() - start_time
    print(f"\n{'='*80}")
    print("PROCESSING COMPLETE")
    print(f"{'='*80}")
    print(f"Time elapsed: {elapsed:.1f} seconds ({elapsed/60:.1f} minutes)")
    print(f"\nOutput files saved to:")
    print(f"  - Patient registries: {OUTPUT_DIR}/")
    print(f"  - Patient notes: {NOTES_DIR}/")
    print(f"\n[OK] All done!")


if __name__ == "__main__":
    main()

