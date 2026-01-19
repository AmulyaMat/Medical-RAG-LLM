"""
build_faiss_index.py
--------------------
Reads medication registry parquet, chunks note_text directly from each row, generates embeddings,
and builds a FAISS vector index for RAG.

Inputs
------
- all_patients_combined.parquet in REGISTRY_PATH
  Required columns:
    - row_id                (unique identifier for each row)
    - patient_id            (string)
    - note_id               (string)
    - note_date             (YYYY-MM-DD)
    - note_text             (full text to be chunked)
    - medication            (medication name)
    - seizure_status        (positive/negative/unknown)
    - seizure_symptoms      (pipe-separated symptoms or None)

Outputs
-------
- OUTPUT_DIR/faiss.index                   (FAISS index)
- OUTPUT_DIR/faiss_chunk_metadata.parquet  (row-per-chunk metadata aligned to the index vectors)
- OUTPUT_DIR/index_info.json               (model name, dim, chunk settings, counts)

Chunking
--------
We use simple, robust, tokenizer-free chunking:
- target_chars: 1200 (≈ 400–800 tokens), overlap_chars: 200
- split at paragraph boundaries when possible, otherwise hard-wrap
Each chunk links back to its source row_id from the registry.

You can safely change CHUNK_* constants to fit your corpus.

Performance Optimizations Applied
----------------------------------
1. ✅ itertuples() instead of iterrows() - 5-10x faster DataFrame iteration
2. ✅ Cross-row batching - Chunk all texts first, then batch embed (better GPU utilization)
3. ✅ FP16 inference - 2x speedup with model.half() on CUDA devices
4. ✅ Excel export disabled by default - Saves 30-60 seconds
5. ✅ Larger batch size (128 vs 64) - Better GPU throughput
6. ✅ Memory efficiency - No intermediate vector list accumulation
"""

import os
import json
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from collections import defaultdict

import faiss
import pandas as pd
from tqdm import tqdm
from datetime import datetime, timezone
import torch
import numpy as np
from transformers import AutoTokenizer, AutoModel

# Import file paths from config.py
from config import REGISTRY_PATH, OUTPUT_DIR, validate_paths

# -----------------
# CONFIG
# -----------------
# Clinical BERT model for medical/EHR text embeddings
EMBEDDING_MODEL = "emilyalsentzer/Bio_ClinicalBERT"   # 768-dim

# Chunking params (character-based)
CHUNK_TARGET_CHARS  = 1200
CHUNK_OVERLAP_CHARS = 200
MIN_CHUNK_CHARS     = 200

# Optional: skip very small/empty texts
MIN_TEXT_CHARS      = 50

# Performance optimizations
USE_FP16 = True  # Enable FP16 inference for 2x speedup (requires CUDA)
EMBEDDING_BATCH_SIZE = 128  # Larger batches for cross-row embedding (increased from 64)

# Excel Export
EXPORT_TO_EXCEL = False  # Set to True to export metadata to Excel file for visualization (saves ~30-60s)

# Check if openpyxl is available for Excel export
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    if EXPORT_TO_EXCEL:
        print("[WARN] openpyxl not installed. Excel export disabled.")
        print("       Install with: pip install openpyxl")
        EXPORT_TO_EXCEL = False

# -----------------
# Utility functions
# -----------------
def load_registry(path: str) -> pd.DataFrame:
    """
    Load medication registry parquet.
    Each row represents one medication record with note_text.
    Each row will be chunked separately and linked via row_id.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Registry not found: {path}")
    
    df = pd.read_parquet(p)
    
    # Validate required columns
    required = ["row_id", "patient_id", "note_id", "note_date", "note_text", "medication"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Registry missing required columns: {missing}")
    
    print(f"[INFO] Loaded {len(df)} medication records")
    return df




def split_paragraphs(text: str) -> List[str]:
    """Split by blank lines; keep non-empty parts."""
    parts = [p.strip() for p in text.split("\n\n")]
    return [p for p in parts if p]


def split_long_paragraph(para: str, target_chars: int, overlap_chars: int) -> List[str]:
    """
    Split a single long paragraph into multiple chunks with overlap.
    Used when a paragraph exceeds target_chars.
    """
    if len(para) <= target_chars:
        return [para]
    
    sub_chunks = []
    i = 0
    while i < len(para):
        end = min(i + target_chars, len(para))
        sub_chunk = para[i:end]
        if len(sub_chunk) >= MIN_CHUNK_CHARS:
            sub_chunks.append(sub_chunk)
        # Move forward with overlap
        i = end - overlap_chars if end - overlap_chars > i else end
        if i >= len(para):
            break
    
    return sub_chunks


def greedy_paragraph_chunk(paragraphs: List[str],
                           target_chars: int,
                           overlap_chars: int) -> List[str]:
    """
    Build chunks by aggregating paragraphs up to ~target_chars.
    If a single paragraph exceeds target_chars, split it into sub-chunks with overlap.
    Overlap is implemented by reusing tail of previous chunk when starting next.
    """
    chunks: List[str] = []
    buf = ""
    for para in paragraphs:
        # Check if this single paragraph is too long
        if len(para) > target_chars:
            # Flush current buffer first
            if buf:
                chunks.append(buf)
                buf = ""
            
            # Split the long paragraph into sub-chunks
            sub_chunks = split_long_paragraph(para, target_chars, overlap_chars)
            chunks.extend(sub_chunks)
            continue
        
        # Normal case: paragraph fits within target
        # Add paragraph with a blank line separator for readability
        candidate = (buf + ("\n\n" if buf else "") + para).strip()
        if len(candidate) <= target_chars:
            buf = candidate
        else:
            # flush current buffer if it's not empty
            if buf:
                chunks.append(buf)
            # start new buffer with para (para is small enough to fit)
            buf = para

    if buf:
        chunks.append(buf)

    # Add overlaps (character-based) between consecutive chunks
    if overlap_chars > 0 and len(chunks) > 1:
        overlapped: List[str] = []
        prev_tail = ""
        for i, ch in enumerate(chunks):
            if i == 0:
                overlapped.append(ch)
                prev_tail = ch[-overlap_chars:] if len(ch) > overlap_chars else ch
            else:
                combined = (prev_tail + "\n\n" + ch).strip()
                overlapped.append(combined)
                prev_tail = ch[-overlap_chars:] if len(ch) > overlap_chars else ch
        chunks = overlapped

    # Final small-chunk filter
    chunks = [c for c in chunks if len(c) >= MIN_CHUNK_CHARS]
    return chunks


def chunk_text(text: str,
               target_chars: int = CHUNK_TARGET_CHARS,
               overlap_chars: int = CHUNK_OVERLAP_CHARS) -> List[str]:
    """
    Paragraph-aware chunking with fallback to hard wraps if there are no blanks.
    """
    if not text or len(text) < MIN_TEXT_CHARS:
        return []

    paras = split_paragraphs(text)
    if not paras:
        # Fallback: hard-wrap by characters
        s = text.strip()
        chunks = []
        i = 0
        while i < len(s):
            end = min(i + target_chars, len(s))
            chunk = s[i:end]
            if len(chunk) >= MIN_CHUNK_CHARS:
                chunks.append(chunk)
            i = end - overlap_chars if end - overlap_chars > i else end
        return chunks

    return greedy_paragraph_chunk(paras, target_chars, overlap_chars)


def normalize_vectors(vecs):
    """
    L2-normalize vectors so that FAISS inner-product ≈ cosine similarity.
    """
    norms = np.linalg.norm(vecs, axis=1, keepdims=True) + 1e-12
    return vecs / norms


def ensure_dir(path: str):
    Path(path).mkdir(parents=True, exist_ok=True)


def export_to_excel(df: pd.DataFrame, output_path: Path, sheet_name: str = "Chunk Metadata") -> bool:
    """
    Export DataFrame to Excel with auto-formatted columns.
    
    Args:
        df: DataFrame to export
        output_path: Path to save Excel file (e.g., "faiss_chunk_metadata.xlsx")
        sheet_name: Name of the Excel sheet
    
    Returns:
        True if successful, False otherwise
    """
    if not EXPORT_TO_EXCEL or not HAS_OPENPYXL:
        return False
    
    if df.empty:
        print(f"[SKIP] No data to export to Excel")
        return False
    
    try:
        # Create Excel file
        excel_path = output_path.with_suffix('.xlsx')
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Auto-adjust column widths
            from openpyxl.utils import get_column_letter
            for col_idx, column in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                # Get header name for special handling
                header_cell = worksheet.cell(row=1, column=col_idx)
                header_name = str(header_cell.value).lower() if header_cell.value else ""
                
                for cell in column:
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                
                # Set width (cap at 50 for readability, except for text columns)
                # chunk_text_full and chunk_preview can be wider
                if 'chunk_text' in header_name:
                    adjusted_width = min(max_length + 2, 100)
                else:
                    adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return True
    except Exception as e:
        print(f"[WARN] Excel export failed: {e}")
        return False


def encode_texts_with_clinicalbert(texts: List[str], tokenizer, model, device, batch_size: int = 32) -> np.ndarray:
    """
    Encode texts using Bio_ClinicalBERT with mean pooling.
    Supports FP16 inference for 2x speedup when model is in half precision.
    
    Args:
        texts: List of text strings to encode
        tokenizer: HuggingFace tokenizer
        model: HuggingFace model (may be FP16 or FP32)
        device: torch device (cuda or cpu)
        batch_size: Batch size for encoding
        
    Returns:
        numpy array of embeddings (num_texts x embedding_dim), always float32
    """
    all_embeddings = []
    is_fp16 = next(model.parameters()).dtype == torch.float16
    
    # Process in batches with progress bar
    for i in tqdm(range(0, len(texts), batch_size), desc="Embedding batches", leave=False):
        batch_texts = texts[i:i + batch_size]
        
        # Tokenize
        encoded_input = tokenizer(
            batch_texts,
            padding=True,
            truncation=True,
            max_length=512,
            return_tensors='pt'
        )
        
        # Move to device
        encoded_input = {k: v.to(device) for k, v in encoded_input.items()}
        
        # Get model outputs
        with torch.no_grad():
            model_output = model(**encoded_input)
        
        # Mean pooling over tokens (excluding padding)
        token_embeddings = model_output.last_hidden_state  # (batch_size, seq_len, hidden_size)
        attention_mask = encoded_input['attention_mask']   # (batch_size, seq_len)
        
        # Expand attention mask for broadcasting
        # Convert to same dtype as embeddings for FP16 compatibility
        attention_mask_expanded = attention_mask.unsqueeze(-1).expand(token_embeddings.size())
        if is_fp16:
            attention_mask_expanded = attention_mask_expanded.half()
        else:
            attention_mask_expanded = attention_mask_expanded.float()
        
        # Sum embeddings weighted by attention mask
        sum_embeddings = torch.sum(token_embeddings * attention_mask_expanded, dim=1)
        
        # Sum attention mask to get count of real tokens
        sum_mask = torch.clamp(attention_mask_expanded.sum(dim=1), min=1e-9)
        
        # Mean pooling
        mean_pooled = sum_embeddings / sum_mask
        
        # Convert to float32 for numpy (even if model is FP16)
        if is_fp16:
            mean_pooled = mean_pooled.float()
        
        # Convert to numpy and add to results
        all_embeddings.append(mean_pooled.cpu().numpy())
    
    # Concatenate all batches
    embeddings = np.vstack(all_embeddings).astype('float32')
    
    # L2 normalize for cosine similarity
    norms = np.linalg.norm(embeddings, axis=1, keepdims=True) + 1e-12
    embeddings = embeddings / norms
    
    return embeddings




# -----------------
# Main builder
# -----------------
def main():
    # Validate configuration paths
    validate_paths()
    ensure_dir(OUTPUT_DIR)
    
    print(f"[INFO] Excel export: {'ENABLED' if EXPORT_TO_EXCEL else 'DISABLED'}")

    # Load full medication registry
    df = load_registry(REGISTRY_PATH)
    print(f"[INFO] Loaded {len(df)} medication records from {df['patient_id'].nunique()} patients")
    print(f"[INFO] Unique medications: {df['medication'].nunique()}")

    # Load embedding model
    print(f"\n[INFO] Loading embedding model: {EMBEDDING_MODEL}")
    tokenizer = AutoTokenizer.from_pretrained(EMBEDDING_MODEL)
    model = AutoModel.from_pretrained(EMBEDDING_MODEL)
    
    # Move model to GPU if available, otherwise CPU
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    model.to(device)
    
    # Enable FP16 for 2x speedup if CUDA is available
    if USE_FP16 and device.type == 'cuda':
        model.half()
        print(f"[INFO] FP16 inference: ENABLED (2x speedup)")
    else:
        print(f"[INFO] FP16 inference: DISABLED (CPU mode or USE_FP16=False)")
    
    model.eval()
    
    dim = 768  # Bio_ClinicalBERT dimension
    print(f"[INFO] Embedding dim: {dim}")
    print(f"[INFO] Using device: {device}")

    # PHASE 1: Chunk all texts and collect metadata (optimized with itertuples)
    print(f"\n[INFO] Phase 1: Chunking {len(df)} rows...")
    metarows: List[Dict] = []
    all_chunks: List[str] = []  # Flat list of all chunks for batch embedding
    
    for row in tqdm(df.itertuples(index=False), total=len(df), desc="Chunking rows"):
        # Extract row data (itertuples is 5-10x faster than iterrows)
        row_id = row.row_id
        patient_id = row.patient_id
        note_id = row.note_id
        note_date = row.note_date
        note_text = row.note_text
        medication = row.medication
        
        # Optional fields (use getattr with default for safety)
        seizure_status = getattr(row, "seizure_status", None)
        seizure_symptoms = getattr(row, "seizure_symptoms", None)
        medication_dosage = getattr(row, "medication_dosage", None)
        intake_times_per_day = getattr(row, "intake_times_per_day", None)
        medication_effectiveness = getattr(row, "medication_effectiveness", None)
        
        # Skip if no text
        if not note_text or pd.isna(note_text):
            continue

        # Chunk the note text
        chunks = chunk_text(str(note_text), CHUNK_TARGET_CHARS, CHUNK_OVERLAP_CHARS)
        if not chunks:
            continue

        # Parse symptoms from pipe-separated string
        symptoms_list = []
        if seizure_symptoms and not pd.isna(seizure_symptoms):
            symptoms_list = [s.strip() for s in str(seizure_symptoms).split('|') if s.strip()]

        # Record metadata for each chunk
        for i, chunk in enumerate(chunks):
            chunk_id = f"{patient_id}_{note_date}_{row_id}_chunk_{i}"
            
            metarows.append({
                # Identifiers
                "vector_id": len(all_chunks),  # Position in flat chunk list
                "chunk_id": chunk_id,
                "source_row_id": row_id,
                "patient_id": str(patient_id),
                "note_id": str(note_id),
                "note_date": str(note_date),
                
                # Chunk info
                "chunk_index": i,
                "chunk_char_len": len(chunk),
                "chunk_preview": chunk[:240].replace("\n", " "),
                "chunk_text_full": chunk,  # Store full chunk for easy retrieval
                
                # Medication info from this row
                "medication": medication,
                "medication_dosage": medication_dosage,
                "intake_times_per_day": intake_times_per_day,
                "medication_effectiveness": medication_effectiveness,
                
                # Seizure info from this row
                "seizure_status": seizure_status,
                "seizure_symptoms": symptoms_list,  # parsed list
                "has_seizure_info": seizure_status == 'positive',
            })
            
            all_chunks.append(chunk)

    # PHASE 2: Batch embed all chunks at once (much more efficient)
    if len(all_chunks) == 0:
        raise RuntimeError("No chunks generated. Check your registry and note_text column.")
    
    print(f"\n[INFO] Phase 2: Embedding {len(all_chunks)} chunks in batches...")
    X = encode_texts_with_clinicalbert(
        all_chunks, 
        tokenizer, 
        model, 
        device, 
        batch_size=EMBEDDING_BATCH_SIZE
    )

    # Build FAISS index
    print(f"\n[INFO] Phase 3: Building FAISS index...")
    # X is already normalized and in float32 from encode_texts_with_clinicalbert
    # Inner-product index (with normalized vectors = cosine similarity)
    index = faiss.IndexFlatIP(X.shape[1])
    index.add(X)

    # Save index
    index_path = Path(OUTPUT_DIR) / "faiss.index"
    faiss.write_index(index, str(index_path))

    # Create metadata DataFrame
    meta_df = pd.DataFrame(metarows)
    # Correct vector_id now that we know the stacked order
    meta_df["vector_id"] = range(len(meta_df))
    
    # Save metadata
    meta_path = Path(OUTPUT_DIR) / "faiss_chunk_metadata.parquet"
    meta_df.to_parquet(meta_path, index=False)
    
    # Export to Excel if enabled
    excel_exported = False
    if EXPORT_TO_EXCEL:
        excel_exported = export_to_excel(meta_df, meta_path, sheet_name="Chunk Metadata")

    # Calculate statistics
    chunks_with_seizures = meta_df['has_seizure_info'].sum() if 'has_seizure_info' in meta_df.columns else 0
    unique_medications = meta_df['medication'].nunique()
    unique_patients = meta_df['patient_id'].nunique()
    unique_source_rows = meta_df['source_row_id'].nunique()
    
    info = {
        "model": EMBEDDING_MODEL,
        "dim": int(X.shape[1]),
        "num_vectors": int(X.shape[0]),
        "num_chunks": int(X.shape[0]),
        "chunk_target_chars": CHUNK_TARGET_CHARS,
        "chunk_overlap_chars": CHUNK_OVERLAP_CHARS,
        "built_utc": datetime.now(timezone.utc).isoformat(),
        "registry_path": REGISTRY_PATH,
        "index_path": str(index_path),
        "metadata_path": str(meta_path),
        
        # Registry-level statistics
        "source_registry_rows": len(df),
        "unique_source_rows_chunked": unique_source_rows,
        "unique_patients": unique_patients,
        "unique_medications": unique_medications,
        "chunks_with_seizure_info": int(chunks_with_seizures),
    }
    
    info_path = Path(OUTPUT_DIR) / "index_info.json"
    with open(info_path, "w", encoding="utf-8") as f:
        json.dump(info, f, indent=2)

    # Print summary
    print("\n" + "="*80)
    print("[OK] FAISS Index Built Successfully")
    print("="*80)
    print(f"Vectors indexed:     {X.shape[0]:,}")
    print(f"Embedding dimension: {X.shape[1]}")
    print(f"Source rows:         {len(df):,}")
    print(f"Unique patients:     {unique_patients}")
    print(f"Unique medications:  {unique_medications}")
    print(f"Chunks with seizure: {chunks_with_seizures:,} ({chunks_with_seizures/len(meta_df)*100:.1f}%)")
    print(f"\nOutput files:")
    print(f"  - {index_path}")
    print(f"  - {meta_path}")
    if excel_exported:
        excel_path = meta_path.with_suffix('.xlsx')
        print(f"  - {excel_path}")
    print(f"  - {info_path}")
    
    # Show top medications
    if len(meta_df) > 0:
        print(f"\nTop 10 medications by chunk count:")
        top_meds = meta_df['medication'].value_counts().head(10)
        for med, count in top_meds.items():
            print(f"  {count:5,} chunks - {med}")


if __name__ == "__main__":
    main()
